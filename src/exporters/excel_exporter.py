"""
Excel exporter for liquidation documents.

Creates well-formatted Excel workbooks with multiple sheets for different data sections.
"""
import pandas as pd
from pathlib import Path
from typing import List
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.models.liquidation import LiquidationDocument


class ExcelExporter:
    """Export liquidation documents to Excel with formatting."""

    def __init__(self, document: LiquidationDocument):
        self.document = document

    def export(self, output_path: str):
        """
        Export document to Excel file.

        Args:
            output_path: Path where Excel file will be saved
        """
        output_path = Path(output_path)

        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write different sections to different sheets
            self._write_document_info(writer)
            self._write_tribute_records(writer)
            self._write_exercise_summaries(writer)
            self._write_deductions(writer)
            self._write_refunds(writer)

        # Apply formatting
        self._apply_formatting(output_path)

    def _write_document_info(self, writer):
        """Write document header information."""
        info_data = {
            'Campo': [
                'Ejercicio',
                'Mandamiento de Pago',
                'Fecha Mandamiento',
                'Número de Liquidación',
                'Código Entidad',
                'Entidad',
                'Total Registros',
                'Total Líquido',
                'A Liquidar',
                'Código Verificación',
                'Firmado Por',
                'Fecha Firma'
            ],
            'Valor': [
                self.document.ejercicio,
                self.document.mandamiento_pago,
                self.document.fecha_mandamiento.strftime('%d/%m/%Y') if self.document.fecha_mandamiento else '',
                self.document.numero_liquidacion,
                self.document.codigo_entidad,
                self.document.entidad,
                self.document.total_records,
                float(self.document.total_liquido),
                float(self.document.a_liquidar),
                self.document.codigo_verificacion or '',
                self.document.firmado_por or '',
                self.document.fecha_firma.strftime('%d/%m/%Y %H:%M:%S') if self.document.fecha_firma else ''
            ]
        }

        df = pd.DataFrame(info_data)
        df.to_excel(writer, sheet_name='Información', index=False)

    def _write_tribute_records(self, writer):
        """Write tribute records (cobros) table."""
        records_data = []

        for record in self.document.tribute_records:
            records_data.append({
                'Ejercicio': record.ejercicio,
                'Concepto': record.concepto,
                'Clave Contabilidad': record.clave_contabilidad,
                'Clave Recaudación': record.clave_recaudacion,
                'Voluntaria': float(record.voluntaria),
                'Ejecutiva': float(record.ejecutiva),
                'Recargo': float(record.recargo),
                'Diputación Voluntaria': float(record.diputacion_voluntaria),
                'Diputación Ejecutiva': float(record.diputacion_ejecutiva),
                'Diputación Recargo': float(record.diputacion_recargo),
                'Líquido': float(record.liquido)
            })

        df = pd.DataFrame(records_data)
        df.to_excel(writer, sheet_name='Registros de Cobros', index=False)

    def _write_exercise_summaries(self, writer):
        """Write exercise summaries."""
        summaries_data = []

        for summary in self.document.exercise_summaries:
            summaries_data.append({
                'Ejercicio': summary.ejercicio,
                'Voluntaria': float(summary.voluntaria),
                'Ejecutiva': float(summary.ejecutiva),
                'Recargo': float(summary.recargo),
                'Diputación Voluntaria': float(summary.diputacion_voluntaria),
                'Diputación Ejecutiva': float(summary.diputacion_ejecutiva),
                'Diputación Recargo': float(summary.diputacion_recargo),
                'Líquido': float(summary.liquido),
                'Número de Registros': len(summary.records)
            })

        # Add overall totals
        summaries_data.append({
            'Ejercicio': 'TOTAL',
            'Voluntaria': float(self.document.total_voluntaria),
            'Ejecutiva': float(self.document.total_ejecutiva),
            'Recargo': float(self.document.total_recargo),
            'Diputación Voluntaria': float(self.document.total_diputacion_voluntaria),
            'Diputación Ejecutiva': float(self.document.total_diputacion_ejecutiva),
            'Diputación Recargo': float(self.document.total_diputacion_recargo),
            'Líquido': float(self.document.total_liquido),
            'Número de Registros': self.document.total_records
        })

        df = pd.DataFrame(summaries_data)
        df.to_excel(writer, sheet_name='Resumen por Ejercicio', index=False)

    def _write_deductions(self, writer):
        """Write deductions and advance breakdown."""
        if not self.document.deductions:
            return

        ded = self.document.deductions

        deductions_data = {
            'Categoría': [
                'RECAUDACIÓN',
                'Tasa Voluntaria',
                'Tasa Ejecutiva',
                'Tasa Ejecutiva Sin Recargo',
                'Tasa Baja Órgano Gestor Deleg.',
                '',
                'TRIBUTARIA',
                'Tasa Gestión Tributaria',
                'Tasa Gestión Censal',
                'Tasa Gestión Catastral',
                '',
                'MULTAS/SANCIONES',
                'Tasa Sanción Tributaria',
                'Tasa Sanción Recaudación',
                'Tasa Sanción Inspección',
                'Tasa Multas de Tráfico',
                '',
                'OTRAS DEDUCCIONES',
                'Gastos Repercutidos',
                'Anticipos',
                'Intereses por Anticipo',
                'Expedientes Compensación',
                'Expedientes Ingresos Indebidos',
                '',
                'TOTAL DEDUCCIONES'
            ],
            'Importe': [
                '',
                float(ded.tasa_voluntaria),
                float(ded.tasa_ejecutiva),
                float(ded.tasa_ejecutiva_sin_recargo),
                float(ded.tasa_baja_organo_gestor_deleg),
                '',
                '',
                float(ded.tasa_gestion_tributaria),
                float(ded.tasa_gestion_censal),
                float(ded.tasa_gestion_catastral),
                '',
                '',
                float(ded.tasa_sancion_tributaria),
                float(ded.tasa_sancion_recaudacion),
                float(ded.tasa_sancion_inspeccion),
                float(ded.tasa_multas_trafico),
                '',
                '',
                float(ded.gastos_repercutidos),
                float(ded.anticipos),
                float(ded.intereses_por_anticipo),
                float(ded.expedientes_compensacion),
                float(ded.expedientes_ingresos_indebidos),
                '',
                float(ded.total_deducciones)
            ]
        }

        df = pd.DataFrame(deductions_data)
        df.to_excel(writer, sheet_name='Deducciones', index=False)

        # Write advance breakdown if available
        if self.document.advance_breakdown:
            advance_data = []
            for adv in self.document.advance_breakdown:
                advance_data.append({
                    'Ejercicio': adv.ejercicio,
                    'Urbana': float(adv.urbana),
                    'Rústica': float(adv.rustica),
                    'Vehículos': float(adv.vehiculos),
                    'BICE': float(adv.bice),
                    'IAE': float(adv.iae),
                    'Tasas': float(adv.tasas),
                    'Ejecutiva': float(adv.ejecutiva),
                    'Total': float(adv.total)
                })

            df_advance = pd.DataFrame(advance_data)
            df_advance.to_excel(writer, sheet_name='Anticipos', index=False)

    def _write_refunds(self, writer):
        """Write refund records."""
        if not self.document.refund_records:
            return

        refunds_data = []
        for refund in self.document.refund_records:
            refunds_data.append({
                'Nº Expediente': refund.num_expte,
                'Nº Resolución': refund.num_resolucion,
                'Nº Solicitud': refund.num_solic,
                'Total Devolución': float(refund.total_devolucion),
                'Entidad': float(refund.entidad),
                'Diputación': float(refund.diputacion),
                'Intereses': float(refund.intereses),
                'Comp. Trib.': float(refund.comp_trib),
                'A Deducir': float(refund.a_deducir)
            })

        df = pd.DataFrame(refunds_data)
        df.to_excel(writer, sheet_name='Devoluciones', index=False)

        # Write summaries if available
        if self.document.refund_summaries:
            summaries_data = []
            for summary in self.document.refund_summaries:
                summaries_data.append({
                    'Concepto': summary.concepto,
                    'Total Devolución': float(summary.total_devolucion),
                    'Entidad': float(summary.entidad),
                    'Diputación': float(summary.diputacion),
                    'Intereses': float(summary.intereses)
                })

            df_summary = pd.DataFrame(summaries_data)
            df_summary.to_excel(writer, sheet_name='Resumen Devoluciones', index=False)

    def _apply_formatting(self, file_path: Path):
        """Apply formatting to the Excel workbook."""
        wb = load_workbook(file_path)

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_font = Font(bold=True, size=11)

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Format number columns
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    # Apply border to all cells
                    cell.border = border

                    # Format numbers
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')

            # Highlight total rows in "Resumen por Ejercicio"
            if sheet_name == 'Resumen por Ejercicio':
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == 'TOTAL':
                        for cell in row:
                            cell.fill = total_fill
                            cell.font = total_font

            # Freeze first row
            ws.freeze_panes = ws['A2']

        wb.save(file_path)


def export_to_excel(document: LiquidationDocument, output_path: str):
    """
    Convenience function to export a liquidation document to Excel.

    Args:
        document: The LiquidationDocument to export
        output_path: Path where Excel file will be saved
    """
    exporter = ExcelExporter(document)
    exporter.export(output_path)
